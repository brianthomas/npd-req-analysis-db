   
#!/usr/bin/env python3
 
'''
    Module for dealing with processing terms.
    
    Created on June 6, 2017
    
    @author: bathomas
'''
    
# Program to load Excel workbook into postgresql database
DesiredWorksheet = 'RQMTs'

SheetToColMap = {
        'Identity' : None,
        'DocName'  : 'docname',
        'Name'     : None,
        'ReqClass' : 'reqclass',
        'Requirement SubClass' : 'reqsubclass',
        'REQUIREMENT' : 'content',
        };
    
def parse (filename):
    
    # load the workbook from file
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    
    # grab the worksheet we want
    if DesiredWorksheet not in wb:
        raise Exception("Can't load desired worksheet from Excel file, bailing")
    
    ws = wb[DesiredWorksheet]
    
    # parse the column names
    colnames = []
    for row in ws.iter_rows(min_row=1, max_col=6, max_row=1):
        for cell in row:
            #print(cell.value)
            mappedVal = SheetToColMap[cell.value]
            colnames.append(mappedVal)
    
    # print (str(colnames))
    
    data = []
    # parse rows to memory array
    for row in ws.iter_rows(min_row=2, max_col=6):
        datum = {}
        i = 0
        for cell in row:
            colname = colnames[i]
            if colname:
                datum[colnames[i]] = cell.value
            i = i + 1
    
        data.append(datum)
    
    return data
    
    
